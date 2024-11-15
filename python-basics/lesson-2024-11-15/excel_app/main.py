from openpyxl import Workbook, load_workbook

file_name = "class_attendance.xlsx"


def create_wb():
    wb = Workbook()

    sheet = wb.active

    sheet["A1"] = "Names"
    sheet["B1"] = "Attendance"

    sheet["A2"] = "Eyong"
    sheet["B2"] = "Present"

    wb.save(filename=file_name)


def create_sheet():
    wb = load_workbook(filename=file_name)

    sheet2 = wb.create_sheet("Sheet 2")
    sheet3 = wb.create_sheet("Sheet 3", 0)

    sheet3.title = "DCI class"

    wb.save(filename=file_name)


if __name__ == "__main__":
    create_sheet()
